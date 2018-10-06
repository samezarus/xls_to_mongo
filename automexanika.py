#!/usr/bin/env python3
# -*- coding: utf8 -*-

# Модуль распознавания файла от "Автомеханика"

import openpyxl

import pymongo

from bson import json_util # Для записи строки в Mongo (интегрирован в pymongo)

import re

class TAutomexanika:
    # Загрузка ексель файла фирмы Автомеханика в базу Mongo

    mongo_db_name = ""
    mongo_db_collection_name = ""

    def __init__(self, xls_file):
        """Constructor"""
        self.xls_file = xls_file
    #
    def jsonToMongo(self, json_string):
        return json_util.loads(json_string)
    #
    def clearString(self, cellValue):
        value = str(cellValue).strip()
        value = value.replace('"', '')
        value = value.replace(',', '')
        value = value.replace(':', '')
        value = value.replace("\\", "/")  # Замена символа "\" на "/"
        value = re.sub('\t|\n|\r|', '', value)  # Удаление из строки Символ табуляции, новой строки и возврата каретки

        return value
    #
    def getElements(self, string, char):
        result = ""

        if (len(string) > 0):
            string = string + char
            b = 0
            for x in range(0, len(string)):
                if (string[x] == char):
                    result = result + string[b:x] + ","
                    b = x + 1
            result = result[0:len(result) - 1]

        return result
    #
    def xlsToMongo(self):
        xls_document = openpyxl.load_workbook(self.xls_file)
        xls_worksheet = xls_document.worksheets[0] # Назначаем индекс листа, который будит парсится

        mongo_db = pymongo.MongoClient()[self.mongo_db_name]
        mongo_db_collection = mongo_db[self.mongo_db_collection_name]

        shop = "automexanika" # Название фирмы, от которой пришёл прайс-лист
        level = ""
        group = ""

        xls_column_count = xls_worksheet.max_column
        xls_row_count = xls_worksheet.max_row

        for row in range(1, xls_row_count):
            index  = 0

            code = ""
            type_product = ""
            brend = ""
            oem = ""
            oem_number = ""
            oem_all_number = ""
            analog = ""
            name = ""
            new_product = ""
            price = ""
            avaible = ""
            incoming = ""

            for column in range(1, xls_column_count):
                index += 1

                value = self.clearString(xls_worksheet.cell(row, column).value)

                if (index == 1): # Пытаемся получить уровень
                    level_var = value

                if (index == 2): # Пытаемся полчить группу
                    if (value != "0" and value != "None"):
                        if (level_var == value):
                            level = level_var # Точнополучили уровень
                            #print(level+":")
                        else:
                            group = value # Точно получили группу
                            #print("    " + group)
                        continue # Пескакиваем на следующую итерацию, так как ловить болше нечего

                if (index == 3): # Пытаемся получить код
                    code = value

                if (index == 4): # Пытаемся получить тип
                    type_product = value

                if (index == 5): # Пытаемся получить производителя
                    brend = value

                if (index == 6): # Пытаемся получить инициалы детали у производителя
                    oem = value

                if (index == 7): # Пытаемся получить номер детали у производителя
                    oem_number = value

                if (index == 8): # Пытаемся получить все номера детали у производителя
                    oem_all_number = self.getElements(value, "/")

                if (index == 9): # Пытаемся получить номера совместимых деталей
                    analog = self.getElements(value, "/")

                if (index == 10): # Пытаемся получить наименование
                    name = value

                if (index == 11): # Пытаемся получить я вляется ли товар новинкой
                    new_product = value

                if (index == 12): # Пытаемся получить цену
                    price = value

                if (index == 13): # Пытаемся получить наличие товара на складе
                    avaible = value

                if (index == 14): # Пытаемся получить приход товара на складе
                    incoming = value

            if (code != "None"):

                json_doc =  '"shop":"'            + shop           + '",'  + \
                            '"level":"'           + level          + '",'  + \
                            '"group":"'           + group          + '",'  + \
                            '"code":"'            + code           + '",'  + \
                            '"type_product":"'    + type_product   + '",'  + \
                            '"brend":"'           + brend          + '",'  + \
                            '"oem":"'             + oem            + '",'  + \
                            '"oem_number":"'      + oem_number     + '",'  + \
                            '"oem_all_number":["' + oem_all_number + '"],' + \
                            '"analog":["'         + analog         + '"],' + \
                            '"name":"'            + name           + '",'  + \
                            '"new_product":"'     + new_product    + '",'  + \
                            '"price":"'           + price          + '",'  + \
                            '"avaible":"'         + avaible        + '",'  + \
                            '"incoming":"'        + incoming       + '"'

                json_doc = '{' + json_doc + '}'
                print (json_doc)
                mongo_doc = self.jsonToMongo(json_doc)
                mongo_db_collection.insert_one(mongo_doc)
#--------------------------------------------------------------------------------------------------------------






