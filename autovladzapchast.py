#!/usr/bin/env python3
# -*- coding: utf8 -*-

# Модуль распознавания файла от "Авто Влад Запчасть"

import openpyxl

import pymongo

from bson import json_util # Для записи строки в Mongo (интегрирован в pymongo)

import re

class TAutoVladZapchast:
    # Загрузка ексель файла фирмы Автомеханика в базу Mongo

    mongo_db_name = ""
    mongo_db_collection_name = ""

    def __init__(self, xls_file):
        """Constructor"""
        self.xls_file = xls_file

    def getColor(self, xls_cell):
        result = ""

        i = xls_cell.fill.start_color.index
        result = str(i)

        return result

    def clearString(self, cellValue):
        value = str(cellValue).strip()
        value = value.replace('"', '')
        value = value.replace(',', '')
        value = value.replace(':', '')
        value = value.replace("\\", "/")  # Замена символа "\" на "/"
        value = re.sub('\t|\n|\r|', '', value)  # Удаление из строки Символ табуляции, новой строки и возврата каретки

        return value

    def xlsToMongo(self):
        xls_document = openpyxl.load_workbook(self.xls_file)
        xls_worksheet = xls_document.worksheets[0] # Назначаем индекс листа, который будит парсится

        mongo_db = pymongo.MongoClient()[self.mongo_db_name]
        mongo_db_collection = mongo_db[self.mongo_db_collection_name]

        shop = "autovladzapchast" # Название фирмы, от которой пришёл прайс-лист
        level = ""
        group = ""

        xls_column_count = xls_worksheet.max_column
        xls_row_count = xls_worksheet.max_row

        for row in range(1, xls_row_count):
            index = 0

            oem_number = ""
            car_brend = ""
            oem= ""
            analog = ""

            for column in range(2, xls_column_count):
                index += 1

                xls_cell = xls_worksheet.cell(row, column)
                value = self.clearString(xls_cell.value)

                if (index == 1):
                    #print (self.getColor(xls_cell))
                    if (self.getColor(xls_cell) == "8"):
                        level = value
                        continue
                    if (self.getColor(xls_cell) == "22"):
                        group = value
                        continue
                    if (self.getColor(xls_cell) == "00000000"):
                        oem_number = value

                if (index == 2):
                    car_brend = value

                if (index == 3):
                    oem = value

            print (level +" | "+ group +" | "+ oem_number +" | " + car_brend+" | " + oem)
#--------------------------------------------------------------------------------------------------------------






