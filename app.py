#!/usr/bin/env python3
# -*- coding: utf8 -*-

# ⊥ ∫ − ″ ¸ ●

import os

import openpyxl
import xlrd

import pymongo

import json
from pprint import pprint

import re

from bson import json_util # Для записи строки в Mongo (интегрирован в pymongo)

from xlstomongo import TXlsToMongo


def jsonReplase (string):
    if(len(string) > 0):
        # Замена символов учавствующих в синтаксисе json
        string = re.sub('"', '″', string)
        # Удаляем литеры из строки
        string = re.sub('\a|\b|\f|\r|\t|\v|', '', string)

        # если в строке есть литер (переноса строки \n), каждую новую строку обрамляем двойными ковычками и разделяем запятыми
        string = re.sub('\n', '","', string)
        string = string.replace("\\", "/")  # Замена символа "\" на "/"
        if (string.find('"') != -1):
            string = '"'+string+'"'
        #
    return string
#
def jsonToMongo(json_string):
    return json_util.loads(json_string)
# --------------------------------------------------------------------------------------------
xlsDir = '/home/sameza/1/'


XlsToMongo = TXlsToMongo()
XlsToMongo.mongo_db_name = 'aridan'

#dirItems = os.listdir(xlsDir)
#for item in dirItems:
#    print (os.listdir(xlsDir+item))






xlsfile = '/home/sameza/1/asva/asva.xls'
conffile = '/home/sameza/1/asva/conf.txt'
if (os.path.exists(xlsfile)):
    XlsToMongo.xls_file = xlsfile
    XlsToMongo.conf_file = conffile
    XlsToMongo.xls_worksheet_namber = 0
    XlsToMongo.mongo_db_collection_name = 'all'
    XlsToMongo.xlsToMongo()


#f = open('/home/sameza/1/conf.txt')
#data = json.load(f)
#print (data)
#print (len(data))
#print (data[0])
#for item in data:
#    print (item)



#mongo_db = pymongo.MongoClient()["aridan"]
#mongo_db_collection = mongo_db["all"]
#mongo_db_collection.remove({})

#file = "/home/sameza/1/automexanika.xlsx"

#Automexanika = automexanika.TAutomexanika(file)
#Automexanika.mongo_db_name = "aridan"
#Automexanika.mongo_db_collection_name = "all"
#Automexanika.xlsToMongo()

#file = "/home/sameza/1/autovladzapchast.xlsx"

#AutoVladZapchast = autovladzapchast.TAutoVladZapchast(file)
#AutoVladZapchast.mongo_db_name = "aridan"
#AutoVladZapchast.mongo_db_collection_name = "all"
#AutoVladZapchast.xlsToMongo()





#xls_document  = openpyxl.load_workbook(file)
#xls_worksheet = xls_document.worksheets[0] # Назначаем индекс листа, который будит парсится

#xls_column_count = xls_worksheet.max_column
#xls_row_count    = xls_worksheet.max_row

#for row in range(1, xls_row_count):
#    json_doc = ''
#    for column in range(1, xls_column_count):
#        field = 'Fileld'+str(column)
#        xls_cell = xls_worksheet.cell(row, column)
#        value = str(xls_cell.value).strip()

#        if (len(value) > 0):

#            value = jsonReplase(value)

#            if (value[0] == '"' and value[len(value) -1] == '"'):
#                value = '['+value+']'
#            else:
#                value = '"'+value+'"'
#        else:
#            value = '""'
#        json_doc = json_doc + '"' + field + '":' + value + ','

#    json_doc = json_doc[0:len(json_doc) - 1]
#    json_doc = r'{' + json_doc + '}'
#    print (json_doc)
#    mongo_doc = jsonToMongo(json_doc)
#    mongo_db_collection.insert_one(mongo_doc)