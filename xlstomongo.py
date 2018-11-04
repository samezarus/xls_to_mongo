#!/usr/bin/env python3
# -*- coding: utf8 -*-

import os

import openpyxl
import xlrd

import pymongo

from bson import json_util # Для записи строки в Mongo (интегрирован в pymongo)

import re

import config

class TXlsToMongo:
    # Загрузка ексель файла фирмы Автомеханика в базу Mongo

    mongo_db_name            = ''
    mongo_db_collection_name = ''

    shop = '' # Название фирмы, от которой пришёл прайс-лист

    xls_file             = '' # эксель файл
    xls_worksheet_namber = 0  # Номер вкладки в эксель файле
    xls_start_row        = 0  # Номер строки с которой начинать парсить эксель файл
    xls_start_column     = 0  # Номер столбца/колонки с которо(го)(й) начинать парсить эксель файл

    conf_file = '' # файл правил обработки эксель файл
    conf_list = [] # список команд из conf_file
    conf_pars = [] # список соманд распарсеных команд из conf_list

    def __init__(self):
        """Constructor"""
        xyz = ''
    #
    def jsonToMongo(self, json_string):
        return json_util.loads(json_string)
    #
    def jsonReplase(self, string):
        if len(string) > 0:
            string = re.sub('"', '″', string)
            string = re.sub('\a|\b|\f|\r|\t|\v|', '', string) # Удаляем литеры из строки
            string = re.sub('\n', '●', string) # если встретели перенос строки, то заменяем его ●
            string = string.replace('\\', '/')  # Замена символа "\" на "/"
        return string
    #
    def getColor(self, xls_cell):
        result = ''

        i = xls_cell.fill.start_color.index
        result = str(i)
        return result
    #
    def loadConfig(self):
        self.conf_list.clear()

        if os.path.exists(self.conf_file):
            file = open(self.conf_file, 'r')
            for line in file:
                if line[0] != '#' and len(line) != 0 :
                    self.conf_list.append(line)
            file.close()
    #
    def getConfig(self):
        self.conf_pars.clear()

        if len(self.conf_list) > 0:
            for item in self.conf_list:
                oneParamStruct = config.getCommandParam(item)
                if oneParamStruct.command == 'shop':
                    self.shop = oneParamStruct.param
                    continue
                if oneParamStruct.command == 'sheetindex':
                    self.xls_worksheet_namber = int(oneParamStruct.param)
                    continue
                if oneParamStruct.command == 'startrow':
                    self.xls_start_row = int(oneParamStruct.param)
                    continue
                if oneParamStruct.command == 'startcolumn':
                    self.xls_start_column = int(oneParamStruct.param)
                    continue

                self.conf_pars.append(config.parsCommandParam(item))
    #
    def xlsToMongo(self):
        if os.path.exists(self.xls_file):
            self.loadConfig()
            self.getConfig()

            mongo_db            = pymongo.MongoClient()[self.mongo_db_name]
            mongo_db_collection = mongo_db[self.mongo_db_collection_name]

            extension = os.path.splitext(self.xls_file)[1]
            if extension == '.xlsx':
                xls_document = openpyxl.load_workbook(self.xls_file)
                xls_worksheet = xls_document.worksheets[self.xls_worksheet_namber] # Назначаем индекс листа, который будит парсится
                xls_column_count = xls_worksheet.max_column
                xls_row_count = xls_worksheet.max_row
            else: # .xls
                xls_document = xlrd.open_workbook(self.xls_file)
                xls_worksheet = xls_document.sheet_by_index(self.xls_worksheet_namber)
                xls_column_count = xls_worksheet.ncols
                xls_row_count = xls_worksheet.nrows

            mongo_db_collection.remove({}) # Очищаем коллекцию(таблицу)

            recordsList = [] # список из значений записи/строки екселя
            json_doc = ''

            #for row in range(0, xls_row_count):
            for row in range(self.xls_start_row, 3):
                json_doc = ''
                for column in range(self.xls_start_column, xls_column_count):
                    field = 'fileld' + str(column)

                    xls_cell = xls_worksheet.cell(row, column)

                    value = str(xls_cell.value).strip()
                    value = self.jsonReplase(value)

                    recordsList.append(value)

                    #if len(value) > 0:
                        #value = '"' + value + '"'
                    #else:
                        #value = '""'
                    #json_doc = json_doc + '"' + field + '":' + value + ','

                if len(recordsList) > 0:
                    if len(self.conf_list) > 0: # если в файле конфигурации есть инструкции
                        for item in self.conf_pars:
                            if len(item) > 0:
                                structCreateName = ''
                                structCreateType = ''
                                structCreateMath = ''
                                structCreateFrom = ''
                                structCreateSplit = ''
                                subResult = ''
                                for subitem in item:
                                    #print (subitem.command + ' : ' + subitem.param)
                                    if subitem.command == 'create':
                                        structCreateName = subitem.param
                                    if subitem.command == 'type':
                                        structCreateType = subitem.param
                                    if subitem.command == 'math':
                                        structCreateMath = subitem.param
                                    if subitem.command == 'split':
                                        structCreateSplit = subitem.param
                                    if subitem.command == 'from':
                                        structCreateFrom = subitem.param

                                fromArray = config.getFrom(structCreateFrom)
                                if len(fromArray) > 0:
                                    for i in range(0, len(fromArray)):
                                        index = int(fromArray[i])
                                        if structCreateType == 'str':
                                            subResult = subResult + recordsList[index] + ' '
                                        if structCreateType == 'float':
                                            getMath = config.getMath(structCreateMath)
                                            if getMath.command == '*':
                                                floatResult = float(recordsList[index]) * float(getMath.param)
                                                subResult = subResult + str(floatResult)  + ' '
                                            if getMath.command == '/':
                                                floatResult = float(recordsList[index]) / float(getMath.param)
                                                subResult = subResult + str(floatResult)  + ' '
                                            if getMath.command == '+':
                                                floatResult = float(recordsList[index]) + float(getMath.param)
                                                subResult = subResult + str(floatResult)  + ' '
                                            if getMath.command == '-':
                                                floatResult = float(recordsList[index]) - float(getMath.param)
                                                subResult = subResult + str(floatResult)  + ' '

                                subResult = '"'+structCreateName+'" = "'+subResult+'"'
                                print(subResult)

                            #print('---------')



                    #else:
                        #json_doc = json_doc[0:len(json_doc) - 1]
                        #json_doc = r'{' + json_doc + '}'
                        #print(json_doc)

                recordsList.clear()

                #json_doc = json_doc[0:len(json_doc) - 1]
                #json_doc = r'{' + json_doc + '}'
                #print(json_doc)
                print('-----------')
                #mongo_doc = self.jsonToMongo(json_doc)
                #mongo_db_collection.insert_one(mongo_doc)


#--------------------------------------------------------------------------------------------------------------






